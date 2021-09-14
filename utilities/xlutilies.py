from openpyxl import Workbook,load_workbook

def getRowcount(filename):
    wb = load_workbook(filename)
    ws = wb.active
    return ws.max_row

def getColcount(filename):
    wb = load_workbook(filename)
    ws = wb.active
    return ws.max_col

def readdata(filename,rownum,columnnum):
    wb=load_workbook(filename)
    ws=wb.active
    return ws.cell(row=rownum,column=columnnum).value

def writedata(filename,rownum,columnnum,data):
    wb =load_workbook(filename)
    ws = wb.active
    ws.cell(row=rownum,column=columnnum).value=data
    wb.save(filename)